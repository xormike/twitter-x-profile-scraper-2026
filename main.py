"""
OSM Kleinort-Extraktor
Vollautomatische Pipeline: Download → Ausschneiden → Konvertieren → Analyse → XLSX
"""

import os
import sys
import subprocess
import urllib.request
import shutil
from pathlib import Path

# UTF-8 Ausgabe erzwingen (Windows-Konsole)
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

def _conda_root_aus_diesem_python():
    """
    Leitet den Conda-Root aus dem aktuell laufenden Python-Interpreter ab.
    start.bat ruft Conda-Python direkt auf, daher ist sys.executable bereits korrekt.
    """
    exe = Path(sys.executable).resolve()
    for candidate in [exe.parent, exe.parent.parent]:
        if (candidate / "Scripts" / "conda.exe").exists() or \
           (candidate / "conda-meta").exists():
            return candidate
    return exe.parent

_CONDA_ROOT = _conda_root_aus_diesem_python()

# ── Abhängigkeiten prüfen (kein Auto-Install) ────────────────────────────────

def _check_import(name):
    try:
        __import__(name)
        return True
    except ImportError:
        return False

print("Pruefe Abhaengigkeiten...")
_FEHLENDE = []
for pkg, imp in [("osmnx", "osmnx"), ("pandas", "pandas"), ("openpyxl", "openpyxl"), ("tqdm", "tqdm")]:
    if _check_import(imp):
        print(f"  OK     {pkg}")
    else:
        print(f"  FEHLT  {pkg}")
        _FEHLENDE.append(pkg)

if _FEHLENDE:
    print()
    print("=" * 60)
    print("Fehlende Pakete werden jetzt automatisch installiert...")
    print("=" * 60)
    print()

    conda_pakete = [p for p in _FEHLENDE if p in ("osmnx",)]
    pip_pakete   = [p for p in _FEHLENDE if p not in conda_pakete]

    fehler = False

    if conda_pakete:
        print(f"  conda install -c conda-forge {' '.join(conda_pakete)}")
        conda_exe = shutil.which("conda") or str(Path(sys.executable).parent / "Scripts" / "conda.exe")
        r = subprocess.run(
            [conda_exe, "install", "-c", "conda-forge", "--solver=classic", "-y"] + conda_pakete,
            text=True
        )
        if r.returncode != 0:
            print("  FEHLER beim conda-Install.")
            fehler = True

    if pip_pakete:
        print(f"  pip install {' '.join(pip_pakete)}")
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet"] + pip_pakete,
            text=True
        )
        if r.returncode != 0:
            print("  FEHLER beim pip-Install.")
            fehler = True

    if fehler:
        print()
        print("Automatische Installation fehlgeschlagen.")
        print("Bitte manuell in der Anaconda Prompt (als Administrator) ausfuehren:")
        if conda_pakete:
            print(f"  conda install -c conda-forge --solver=classic {' '.join(conda_pakete)}")
        if pip_pakete:
            print(f"  pip install {' '.join(pip_pakete)}")
        input("\nEnter zum Beenden...")
        sys.exit(1)
    else:
        print()
        print("Installation abgeschlossen. Pakete werden geladen...")
        print()

print()

import osmnx as ox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ─────────────────────────────────────────────
# Geofabrik-Regionen
# ─────────────────────────────────────────────

REGIONEN = {
    # Deutschland
    "Bayern":               "https://download.geofabrik.de/europe/germany/bavaria-latest.osm.pbf",
    "Baden-Württemberg":    "https://download.geofabrik.de/europe/germany/baden-wuerttemberg-latest.osm.pbf",
    "Berlin":               "https://download.geofabrik.de/europe/germany/berlin-latest.osm.pbf",
    "Brandenburg":          "https://download.geofabrik.de/europe/germany/brandenburg-latest.osm.pbf",
    "Bremen":               "https://download.geofabrik.de/europe/germany/bremen-latest.osm.pbf",
    "Hamburg":              "https://download.geofabrik.de/europe/germany/hamburg-latest.osm.pbf",
    "Hessen":               "https://download.geofabrik.de/europe/germany/hessen-latest.osm.pbf",
    "Mecklenburg-Vorpommern": "https://download.geofabrik.de/europe/germany/mecklenburg-vorpommern-latest.osm.pbf",
    "Niedersachsen":        "https://download.geofabrik.de/europe/germany/niedersachsen-latest.osm.pbf",
    "Nordrhein-Westfalen":  "https://download.geofabrik.de/europe/germany/nordrhein-westfalen-latest.osm.pbf",
    "Rheinland-Pfalz":      "https://download.geofabrik.de/europe/germany/rheinland-pfalz-latest.osm.pbf",
    "Saarland":             "https://download.geofabrik.de/europe/germany/saarland-latest.osm.pbf",
    "Sachsen":              "https://download.geofabrik.de/europe/germany/sachsen-latest.osm.pbf",
    "Sachsen-Anhalt":       "https://download.geofabrik.de/europe/germany/sachsen-anhalt-latest.osm.pbf",
    "Schleswig-Holstein":   "https://download.geofabrik.de/europe/germany/schleswig-holstein-latest.osm.pbf",
    "Thüringen":            "https://download.geofabrik.de/europe/germany/thueringen-latest.osm.pbf",
    # Österreich
    "Österreich":           "https://download.geofabrik.de/europe/austria-latest.osm.pbf",
    # Schweiz
    "Schweiz":              "https://download.geofabrik.de/europe/switzerland-latest.osm.pbf",
    # Weitere Europa
    "Italien":              "https://download.geofabrik.de/europe/italy-latest.osm.pbf",
    "Frankreich":           "https://download.geofabrik.de/europe/france-latest.osm.pbf",
    "Spanien":              "https://download.geofabrik.de/europe/spain-latest.osm.pbf",
    "Polen":                "https://download.geofabrik.de/europe/poland-latest.osm.pbf",
    "Tschechien":           "https://download.geofabrik.de/europe/czech-republic-latest.osm.pbf",
    # Eigene URL
    "Eigene URL eingeben":  None,
}

# Ortstyp-Übersetzungen
ORTSTYP_DE = {
    "hamlet":            "Weiler",
    "village":           "Dorf",
    "suburb":            "Stadtteil",
    "isolated_dwelling": "Einzelgehöft",
    "locality":          "Flurname",
    "town":              "Kleinstadt",
    "city":              "Stadt",
    "municipality":      "Gemeinde",
    "district":          "Bezirk",
    "borough":           "Stadtbezirk",
    "quarter":           "Stadtviertel",
    "neighbourhood":     "Nachbarschaft",
}

WORK_DIR = Path(__file__).parent / "daten"
WORK_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
# Hilfsfunktionen
# ─────────────────────────────────────────────

def trennlinie(zeichen="─", breite=60):
    print(zeichen * breite)

def titel(text):
    trennlinie()
    print(f"  {text}")
    trennlinie()

def osmium_pfad():
    """
    Sucht osmium relativ zum aktiven Python-Interpreter.
    sys.executable zeigt auf das Conda-Python, osmium liegt
    im selben Environment unter Library/bin/ (Windows).
    """
    kandidaten = []

    # 1. shutil.which (falls osmium im PATH ist)
    w = shutil.which("osmium")
    if w:
        kandidaten.append(Path(w))

    # 2. Relativ zu sys.executable — robusteste Methode
    python_dir = Path(sys.executable).resolve().parent
    kandidaten += [
        python_dir / "Library" / "bin" / "osmium.exe",
        python_dir / "bin" / "osmium",
        python_dir / "osmium.exe",
        python_dir.parent / "Library" / "bin" / "osmium.exe",
        python_dir.parent / "bin" / "osmium",
    ]

    # 3. Conda-Root Fallback
    if _CONDA_ROOT:
        root = Path(_CONDA_ROOT)
        kandidaten += [
            root / "Library" / "bin" / "osmium.exe",
            root / "bin" / "osmium.exe",
        ]

    for k in kandidaten:
        if Path(k).exists():
            return str(Path(k).resolve())
    return None


def download_pbf(url, ziel_pfad):
    """Lädt PBF-Datei mit Fortschrittsbalken herunter. Wiederholt bei Abbruch."""
    import urllib.request
    import threading
    MIN_GROESSE_MB = 1
    TIMEOUT_SEKUNDEN = 30  # Abbruch wenn 30s kein neues Datenpaket

    # Erwartete Dateigröße per HEAD-Request abfragen
    erwartete_mb = None
    try:
        req = urllib.request.Request(url, method="HEAD")
        with urllib.request.urlopen(req, timeout=10) as resp:
            content_length = resp.headers.get("Content-Length")
            if content_length:
                erwartete_mb = int(content_length) / (1024 * 1024)
                print(f"  Erwartete Groesse: {erwartete_mb:.0f} MB")
    except Exception:
        pass  # HEAD fehlgeschlagen — kein Problem, weiter ohne Größeninfo

    class FortschrittsHaken(tqdm):
        def update_to(self, b=1, bsize=1, tsize=None):
            if tsize is not None:
                self.total = tsize
            self.update(b * bsize - self.n)

    def _download_mit_timeout(url, ziel_pfad, haken):
        """Lädt herunter und wirft TimeoutError wenn zu lange kein Fortschritt."""
        letzter_stand = [0]
        abbruch = [False]
        fehler = [None]

        def _waechter():
            import time
            while not abbruch[0]:
                stand_vorher = letzter_stand[0]
                time.sleep(TIMEOUT_SEKUNDEN)
                if abbruch[0]:
                    break
                if letzter_stand[0] == stand_vorher and not abbruch[0]:
                    fehler[0] = TimeoutError(
                        f"Kein Fortschritt seit {TIMEOUT_SEKUNDEN} Sekunden — Verbindung eingefroren?"
                    )
                    # urllib hat keinen externen Abbruch-Hook, daher Socket schliessen
                    import socket
                    try:
                        socket.setdefaulttimeout(0.001)
                    except Exception:
                        pass
                    break

        waechter = threading.Thread(target=_waechter, daemon=True)
        waechter.start()

        original_update = haken.update_to
        def _update_mit_stand(b=1, bsize=1, tsize=None):
            letzter_stand[0] = b * bsize
            original_update(b, bsize, tsize)
        haken.update_to = _update_mit_stand

        try:
            urllib.request.urlretrieve(url, ziel_pfad, reporthook=haken.update_to)
        except Exception as e:
            if fehler[0]:
                raise fehler[0]
            raise e
        finally:
            abbruch[0] = True
            import socket
            socket.setdefaulttimeout(None)

        if fehler[0]:
            raise fehler[0]

    while True:
        print(f"\nQuelle: {url}")
        print(f"Ziel:   {ziel_pfad}")
        try:
            with FortschrittsHaken(unit='B', unit_scale=True, unit_divisor=1024,
                                   miniters=1, desc="Download") as t:
                _download_mit_timeout(url, ziel_pfad, t)
        except Exception as e:
            print(f"\n  Download unterbrochen: {e}")
            if Path(ziel_pfad).exists():
                Path(ziel_pfad).unlink()
            antwort = input("  Erneut versuchen? (j/n) [j]: ").strip().lower()
            if antwort != "n":
                continue
            raise RuntimeError("Download abgebrochen.")

        groesse_mb = Path(ziel_pfad).stat().st_size / (1024 * 1024)
        # Prüfen: zu klein absolut, oder deutlich kleiner als erwartet
        zu_klein = groesse_mb < MIN_GROESSE_MB
        zu_kurz = erwartete_mb and groesse_mb < erwartete_mb * 0.95
        if zu_klein or zu_kurz:
            if zu_kurz and not zu_klein:
                print(f"\n  FEHLER: Datei unvollstaendig ({groesse_mb:.0f} MB von erwartet {erwartete_mb:.0f} MB).")
            else:
                print(f"\n  FEHLER: Heruntergeladene Datei ist nur {groesse_mb:.2f} MB gross.")
            print("  Der Download war wahrscheinlich unvollstaendig (instabile Verbindung).")
            Path(ziel_pfad).unlink()
            print()
            print("  Optionen:")
            print("  1. Automatisch erneut versuchen")
            print("  2. Datei manuell herunterladen (Browser oeffnen)")
            print("  3. Abbrechen")
            print()
            while True:
                wahl = input("  Auswahl (1/2/3): ").strip()
                if wahl == "1":
                    break
                elif wahl == "2":
                    import webbrowser
                    webbrowser.open(url)
                    print()
                    print("  Datei manuell herunterladen, dann hier fortfahren.")
                    print()
                    print(f"  Empfohlen: Datei in diesen Ordner verschieben:")
                    print(f"  --> {WORK_DIR}")
                    print()
                    input("  Enter druecken wenn der Download abgeschlossen ist...")
                    # Ordner automatisch nach neuen PBF-Dateien durchsuchen (wie Schritt 2)
                    gefunden = pbf_dateien_suchen()
                    if gefunden:
                        print()
                        print("  Folgende PBF-Dateien wurden gefunden:\n")
                        for i, p in enumerate(gefunden, 1):
                            groesse_mb_f = p.stat().st_size / (1024 * 1024)
                            print(f"    {i}. {p.name}  ({groesse_mb_f:.0f} MB)")
                            print(f"       {p.parent}")
                        print()
                        while True:
                            auswahl = input("  Nummer der Datei waehlen: ").strip()
                            try:
                                idx = int(auswahl) - 1
                                if 0 <= idx < len(gefunden):
                                    manuell = gefunden[idx]
                                    groesse_manuell = manuell.stat().st_size / (1024 * 1024)
                                    if groesse_manuell < MIN_GROESSE_MB:
                                        print(f"  Datei zu klein ({groesse_manuell:.2f} MB) — Download unvollstaendig?")
                                        continue
                                    import shutil as _shutil
                                    _shutil.copy2(str(manuell), str(ziel_pfad))
                                    print(f"✅ Datei uebernommen ({groesse_manuell:.0f} MB).")
                                    return
                            except ValueError:
                                pass
                            print("  Ungueltige Auswahl.")
                    else:
                        print("  Keine PBF-Dateien gefunden.")
                        print(f"  Bitte Datei manuell in diesen Ordner legen und erneut starten:")
                        print(f"  --> {WORK_DIR}")
                        raise RuntimeError("Download abgebrochen.")
                    break
                elif wahl == "3":
                    raise RuntimeError("Download abgebrochen.")
                else:
                    print("  Bitte 1, 2 oder 3 eingeben.")
            if wahl == "1":
                continue

        print(f"✅ Download abgeschlossen ({groesse_mb:.0f} MB).")
        break

def bbox_eingabe():
    """Interaktive Bounding-Box-Eingabe mit Validierung."""
    print("\nBounding Box eingeben:")
    print("Format: min_lon,min_lat,max_lon,max_lat")
    print("Beispiel Muenchen: 11.360232,48.061602,11.722837,48.248220\n")

    while True:
        eingabe = input("Bounding Box: ").strip()
        teile = eingabe.replace(";", ",").split(",")
        if len(teile) != 4:
            print("  Fehler: Genau 4 Werte durch Komma trennen.")
            continue
        try:
            min_lon, min_lat, max_lon, max_lat = [float(t.strip()) for t in teile]
        except ValueError:
            print("  Fehler: Nur Zahlen erlaubt.")
            continue
        if min_lon >= max_lon or min_lat >= max_lat:
            print("  Fehler: min-Werte müssen kleiner als max-Werte sein.")
            continue
        if not (-180 <= min_lon <= 180 and -180 <= max_lon <= 180):
            print("  Fehler: Längengrad muss zwischen -180 und 180 liegen.")
            continue
        if not (-90 <= min_lat <= 90 and -90 <= max_lat <= 90):
            print("  Fehler: Breitengrad muss zwischen -90 und 90 liegen.")
            continue
        return min_lon, min_lat, max_lon, max_lat

def extrahieren(osmium, bbox, quell_pbf, ziel_pbf):
    """Schneidet Region aus PBF heraus."""
    min_lon, min_lat, max_lon, max_lat = bbox
    bbox_str = f"{min_lon},{min_lat},{max_lon},{max_lat}"
    print(f"\nExtrahiere Bounding Box: {bbox_str}")
    result = subprocess.run(
        [Path(osmium), "extract", "-b", bbox_str, Path(quell_pbf),
         "-o", Path(ziel_pbf), "--overwrite"],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"osmium extract fehlgeschlagen:\n{result.stderr}")
    print("✅ Extraktion abgeschlossen.")

def konvertieren(osmium, quell_pbf, ziel_osm):
    """Konvertiert PBF → OSM (XML)."""
    print(f"\nKonvertiere {quell_pbf.name} → {ziel_osm.name}")
    print("Hinweis: Dieser Schritt kann einige Minuten dauern und viel RAM benötigen.")
    result = subprocess.run(
        [Path(osmium), "cat", Path(quell_pbf), "-o", Path(ziel_osm), "--overwrite"],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"osmium cat fehlgeschlagen:\n{result.stderr}")
    print("✅ Konvertierung abgeschlossen.")

def analysieren(osm_datei):
    """Extrahiert Kleinorte mit osmnx."""
    print(f"\nAnalysiere {osm_datei.name}...")
    tags = {"place": [
        "hamlet", "village", "suburb", "isolated_dwelling",
        "locality", "town", "quarter", "neighbourhood", "borough"
    ]}
    gdf = ox.features_from_xml(str(osm_datei), tags=tags)

    # Relevante Spalten sammeln
    spalten = ["name", "place"]
    for sp in ["addr:suburb", "addr:city", "addr:municipality",
               "is_in", "is_in:municipality", "addr:district"]:
        if sp in gdf.columns:
            spalten.append(sp)

    df = gdf[spalten].copy()
    df = df.dropna(subset=["name"]).drop_duplicates(subset=["name", "place"])
    df = df.reset_index(drop=True)

    # Koordinaten hinzufügen (Centroid bei Polygonen)
    try:
        geom = gdf.loc[df.index, "geometry"]
        centroids = geom.centroid
        df["lat"] = centroids.y.round(6)
        df["lon"] = centroids.x.round(6)
    except Exception:
        df["lat"] = None
        df["lon"] = None

    print(f"✅ {len(df)} Orte gefunden.")
    return df

def gemeinde_ermitteln(row):
    """Versucht Gemeindezugehörigkeit aus verschiedenen Feldern zu lesen."""
    for sp in ["addr:municipality", "is_in:municipality", "addr:city", "addr:suburb", "is_in"]:
        val = row.get(sp)
        if pd.notna(val) and val:
            # is_in kann mehrere Werte enthalten, erstes nehmen
            if isinstance(val, str) and "," in val:
                return val.split(",")[0].strip()
            return str(val).strip()
    return ""

def xlsx_exportieren(df, pfad, region_name, bbox):
    """Exportiert strukturierte XLSX-Tabelle."""
    print(f"\nErstelle XLSX: {pfad.name}")

    wb = Workbook()

    # ── Blatt 1: Übersicht nach Ortstyp ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Orte nach Typ"

    # Farben
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    SECTION_FILL  = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
    TITLE_FILL    = PatternFill("solid", fgColor="0D2137")

    HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    SECTION_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    TITLE_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    BODY_FONT     = Font(name="Arial", size=10)

    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    def kopfzelle(zelle, text, fill=HEADER_FILL, font=HEADER_FONT):
        zelle.value = text
        zelle.fill = fill
        zelle.font = font
        zelle.alignment = center
        zelle.border = border

    def datenzelle(zelle, text, fill=None, alignierung=left):
        zelle.value = text
        zelle.font = BODY_FONT
        zelle.alignment = alignierung
        zelle.border = border
        if fill:
            zelle.fill = fill

    # Titelzeile
    ws1.merge_cells("A1:E1")
    kopfzelle(ws1["A1"], f"OSM Kleinort-Auswertung — {region_name}", TITLE_FILL, TITLE_FONT)
    ws1.row_dimensions[1].height = 28

    # Metadaten
    ws1.merge_cells("A2:E2")
    bbox_text = f"Bounding Box: {bbox[0]}, {bbox[1]}, {bbox[2]}, {bbox[3]}"
    kopfzelle(ws1["A2"], bbox_text, PatternFill("solid", fgColor="1A3A54"), Font(color="AACCE8", name="Arial", size=9))

    # Spaltenheader
    headers = ["Nr.", "Name", "Ortstyp (DE)", "Ortstyp (OSM)", "Gemeinde"]
    for col, h in enumerate(headers, 1):
        kopfzelle(ws1.cell(3, col), h)
    ws1.row_dimensions[3].height = 22

    # Daten gruppiert nach Ortstyp
    df_sorted = df.copy()
    df_sorted["ortstyp_de"] = df_sorted["place"].map(lambda x: ORTSTYP_DE.get(x, x.capitalize()))
    df_sorted["gemeinde"]   = df_sorted.apply(gemeinde_ermitteln, axis=1)
    df_sorted = df_sorted.sort_values(["ortstyp_de", "name"])

    zeile = 4
    gesamt_nr = 1

    for ortstyp_de, gruppe in df_sorted.groupby("ortstyp_de"):
        # Abschnittsheader
        ws1.merge_cells(f"A{zeile}:E{zeile}")
        kopfzelle(ws1.cell(zeile, 1), f"{ortstyp_de}  ({len(gruppe)} Einträge)", SECTION_FILL, SECTION_FONT)
        ws1.row_dimensions[zeile].height = 20
        zeile += 1

        for i, (_, row) in enumerate(gruppe.iterrows()):
            fill = ALT_FILL if i % 2 == 0 else None
            osm_typ = row.get("place", "")

            datenzelle(ws1.cell(zeile, 1), gesamt_nr, fill, center)
            datenzelle(ws1.cell(zeile, 2), row.get("name", ""), fill)
            datenzelle(ws1.cell(zeile, 3), ORTSTYP_DE.get(osm_typ, osm_typ.capitalize()), fill)
            datenzelle(ws1.cell(zeile, 4), osm_typ, fill)
            datenzelle(ws1.cell(zeile, 5), row.get("gemeinde", ""), fill)

            ws1.row_dimensions[zeile].height = 18
            zeile += 1
            gesamt_nr += 1

    # Summenzeile
    ws1.merge_cells(f"A{zeile}:E{zeile}")
    kopfzelle(ws1.cell(zeile, 1), f"Gesamt: {len(df_sorted)} Orte", HEADER_FILL)
    ws1.row_dimensions[zeile].height = 20

    # Spaltenbreiten
    for col, breite in zip(range(1, 6), [6, 28, 18, 16, 24]):
        ws1.column_dimensions[get_column_letter(col)].width = breite

    # ── Blatt 2: Alphabetische Liste ─────────────────────────────────
    ws2 = wb.create_sheet("Alphabetisch")

    ws2.merge_cells("A1:E1")
    kopfzelle(ws2["A1"], "Alle Orte — alphabetisch", TITLE_FILL, TITLE_FONT)
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        kopfzelle(ws2.cell(2, col), h)

    df_alpha = df_sorted.sort_values("name")
    for i, (_, row) in enumerate(df_alpha.iterrows()):
        z = i + 3
        fill = ALT_FILL if i % 2 == 0 else None
        osm_typ = row.get("place", "")
        datenzelle(ws2.cell(z, 1), i + 1, fill, center)
        datenzelle(ws2.cell(z, 2), row.get("name", ""), fill)
        datenzelle(ws2.cell(z, 3), ORTSTYP_DE.get(osm_typ, osm_typ.capitalize()), fill)
        datenzelle(ws2.cell(z, 4), osm_typ, fill)
        datenzelle(ws2.cell(z, 5), row.get("gemeinde", ""), fill)
        ws2.row_dimensions[z].height = 18

    for col, breite in zip(range(1, 6), [6, 28, 18, 16, 24]):
        ws2.column_dimensions[get_column_letter(col)].width = breite

    # ── Blatt 3: Statistik ────────────────────────────────────────────
    ws3 = wb.create_sheet("Statistik")

    ws3.merge_cells("A1:C1")
    kopfzelle(ws3["A1"], "Statistik nach Ortstyp", TITLE_FILL, TITLE_FONT)
    ws3.row_dimensions[1].height = 28

    for col, h in enumerate(["Ortstyp (DE)", "Ortstyp (OSM)", "Anzahl"], 1):
        kopfzelle(ws3.cell(2, col), h)

    statistik = df_sorted.groupby(["ortstyp_de", "place"]).size().reset_index(name="Anzahl")
    statistik = statistik.sort_values("Anzahl", ascending=False)

    for i, (_, row) in enumerate(statistik.iterrows()):
        z = i + 3
        fill = ALT_FILL if i % 2 == 0 else None
        datenzelle(ws3.cell(z, 1), row["ortstyp_de"], fill)
        datenzelle(ws3.cell(z, 2), row["place"], fill)
        datenzelle(ws3.cell(z, 3), row["Anzahl"], fill, center)
        ws3.row_dimensions[z].height = 18

    # Summe
    sum_zeile = len(statistik) + 3
    ws3.merge_cells(f"A{sum_zeile}:B{sum_zeile}")
    kopfzelle(ws3.cell(sum_zeile, 1), "Gesamt", HEADER_FILL)
    ws3.cell(sum_zeile, 3).value = f"=SUM(C3:C{sum_zeile-1})"
    ws3.cell(sum_zeile, 3).font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    ws3.cell(sum_zeile, 3).fill = HEADER_FILL
    ws3.cell(sum_zeile, 3).alignment = center
    ws3.cell(sum_zeile, 3).border = border

    for col, breite in zip(range(1, 4), [22, 18, 12]):
        ws3.column_dimensions[get_column_letter(col)].width = breite

    wb.save(str(pfad))
    print(f"✅ XLSX gespeichert: {pfad}")

# ─────────────────────────────────────────────
# Hauptprogramm
# ─────────────────────────────────────────────

def pbf_dateien_suchen():
    """Sucht PBF-Dateien im Arbeitsverzeichnis und auf dem gesamten System (gängige Orte)."""
    gefunden = []

    # 1. Im Daten-Unterordner des Skripts
    for p in sorted(WORK_DIR.glob("*.pbf")):
        gefunden.append(p)

    # 2. Im Skript-Verzeichnis selbst
    skript_dir = Path(__file__).parent
    for p in sorted(skript_dir.glob("*.pbf")):
        if p not in gefunden:
            gefunden.append(p)

    # 3. Gängige Download-Orte auf Windows
    extra_orte = [
        Path.home() / "Downloads",
        Path("C:/Users/Public/Downloads"),
        Path("D:/"),
    ]
    for ordner in extra_orte:
        if ordner.exists():
            for p in sorted(ordner.glob("*.pbf")):
                if p not in gefunden:
                    gefunden.append(p)

    return gefunden


def pbf_datei_waehlen(download_url):
    """
    Interaktive Auswahl der PBF-Quelldatei.
    Prüft zuerst ob lokale PBFs vorhanden sind, bietet Auswahl an,
    und lädt als letzten Ausweg von Geofabrik herunter.
    """
    vorhandene = pbf_dateien_suchen()

    if vorhandene:
        print("Folgende PBF-Dateien wurden auf diesem System gefunden:\n")
        for i, p in enumerate(vorhandene, 1):
            groesse_mb = p.stat().st_size / (1024 * 1024)
            print(f"  {i:>3}. {p.name}  ({groesse_mb:.0f} MB)")
            print(f"       {p.parent}")
        print()
        print(f"  {len(vorhandene)+1:>3}. Anderen Pfad manuell eingeben")
        print(f"  {len(vorhandene)+2:>3}. Datei jetzt herunterladen ({download_url.split('/')[-1]})")
        print()

        while True:
            auswahl = input("Nummer eingeben: ").strip()
            try:
                idx = int(auswahl)
                if 1 <= idx <= len(vorhandene):
                    gewählt = vorhandene[idx - 1]
                    print(f"✅ Verwende: {gewählt}")
                    return gewählt
                elif idx == len(vorhandene) + 1:
                    return pbf_pfad_manuell()
                elif idx == len(vorhandene) + 2:
                    return pbf_herunterladen(download_url)
            except ValueError:
                pass
            print("  Ungültige Auswahl.")
    else:
        print("Keine PBF-Dateien im Verzeichnis gefunden.\n")
        print("  1. Pfad manuell eingeben")
        print(f"  2. Datei jetzt herunterladen ({download_url.split('/')[-1]})")
        print()

        while True:
            auswahl = input("Nummer eingeben: ").strip()
            if auswahl == "1":
                return pbf_pfad_manuell()
            elif auswahl == "2":
                return pbf_herunterladen(download_url)
            print("  Bitte 1 oder 2 eingeben.")


def pbf_pfad_manuell():
    """Lässt den User einen vollständigen Pfad zu einer PBF-Datei eingeben."""
    while True:
        pfad_str = input("Vollständiger Pfad zur .pbf-Datei: ").strip().strip('"')
        pfad = Path(pfad_str)
        if not pfad_str:
            print("  Kein Pfad eingegeben.")
            continue
        if not pfad.exists():
            print(f"  Datei nicht gefunden: {pfad}")
            continue
        if pfad.suffix.lower() != ".pbf":
            print("  Warnung: Datei hat keine .pbf-Endung — trotzdem verwenden? (j/n)")
            if input().strip().lower() != "j":
                continue
        print(f"✅ Verwende: {pfad}")
        return pfad


def pbf_herunterladen(url):
    """Lädt die PBF-Datei in den Datenordner herunter und gibt den Pfad zurück."""
    dateiname = url.split("/")[-1]
    ziel = WORK_DIR / dateiname
    if ziel.exists():
        groesse_mb = ziel.stat().st_size / (1024 * 1024)
        print(f"Datei bereits vorhanden: {ziel.name} ({groesse_mb:.0f} MB)")
        neu = input("Neu herunterladen? (j/n) [n]: ").strip().lower()
        if neu != "j":
            print(f"✅ Verwende vorhandene Datei: {ziel}")
            return ziel

    while True:
        try:
            download_pbf(url, ziel)
            return ziel
        except RuntimeError:
            # Download fehlgeschlagen oder abgebrochen — Optionen anbieten
            print()
            print("  Optionen:")
            print("  1. Automatisch erneut versuchen")
            print("  2. Datei manuell herunterladen (Browser oeffnen)")
            print("  3. Abbrechen")
            print()
            while True:
                wahl = input("  Auswahl (1/2/3): ").strip()
                if wahl == "1":
                    break  # Schleife erneut
                elif wahl == "2":
                    import webbrowser
                    webbrowser.open(url)
                    print()
                    print(f"  Empfohlen: Datei in diesen Ordner verschieben:")
                    print(f"  --> {WORK_DIR}")
                    print()
                    input("  Enter druecken wenn der Download abgeschlossen ist...")
                    # Automatisch im daten/-Ordner und Standardorten suchen
                    gefunden = pbf_dateien_suchen()
                    if gefunden:
                        print()
                        print("  Gefundene PBF-Dateien:")
                        for i, p in enumerate(gefunden, 1):
                            mb = p.stat().st_size / (1024 * 1024)
                            print(f"    {i}. {p.name}  ({mb:.0f} MB)  —  {p.parent}")
                        print()
                        while True:
                            nr = input("  Nummer auswaehlen: ").strip()
                            try:
                                nr_idx = int(nr) - 1
                                if 0 <= nr_idx < len(gefunden):
                                    print(f"  ✅ Verwende: {gefunden[nr_idx]}")
                                    return gefunden[nr_idx]
                            except ValueError:
                                pass
                            print("  Ungueltige Auswahl.")
                    else:
                        print("  Keine PBF-Datei gefunden. Bitte Datei manuell in folgenden Ordner legen:")
                        print(f"  --> {WORK_DIR}")
                        input("  Dann Enter druecken...")
                        # Nochmal suchen
                        gefunden2 = pbf_dateien_suchen()
                        if gefunden2:
                            return gefunden2[0]
                        print("  Immer noch keine Datei gefunden. Abbruch.")
                        sys.exit(1)
                elif wahl == "3":
                    print("  Abgebrochen.")
                    sys.exit(0)
                else:
                    print("  Bitte 1, 2 oder 3 eingeben.")


def main():
    titel("OSM Kleinort-Extraktor")
    print("Erstellt eine strukturierte XLSX-Tabelle aus OpenStreetMap-Daten.\n")

    # Osmium prüfen
    osmium = osmium_pfad()
    if not osmium:
        print("FEHLER: osmium nicht gefunden!")
        print("Bitte in Anaconda Prompt ausführen: conda install -c conda-forge osmium-tool")
        input("\nEnter zum Beenden...")
        sys.exit(1)
    print(f"✅ osmium gefunden: {osmium}")

    # ── Region auswählen ──────────────────────────────────────────────
    titel("1/5 · Region auswählen")
    region_namen = list(REGIONEN.keys())
    for i, name in enumerate(region_namen, 1):
        print(f"  {i:>3}. {name}")
    print()

    while True:
        auswahl = input("Nummer eingeben: ").strip()
        try:
            idx = int(auswahl) - 1
            if 0 <= idx < len(region_namen):
                break
        except ValueError:
            pass
        print("  Ungültige Auswahl, bitte Nummer eingeben.")

    region_name = region_namen[idx]
    url = REGIONEN[region_name]

    if url is None:
        url = input("Geofabrik-URL eingeben (.osm.pbf): ").strip()

    # ── PBF-Datei wählen oder herunterladen ──────────────────────────
    titel("2/5 · PBF-Datei")
    pbf_pfad = pbf_datei_waehlen(url)

    # ── Bounding Box ──────────────────────────────────────────────────
    titel("3/5 · Bounding Box")
    print("Die Website zum Zeichnen der Bounding Box wird jetzt geoeffnet.")
    print()
    print("  Anleitung:")
    print("  1. Region auf der Karte einzeichnen")
    print("  2. Unten links das Format auf  CSV  umstellen  (nicht MARC!)")
    print("  3. Die 4 Zahlen kopieren und hier einfuegen")
    print()
    import webbrowser
    import threading
    import time

    def csv_hinweis_popup():
        time.sleep(3)
        if sys.platform == "win32":
            import ctypes
            # MB_ICONINFORMATION | MB_SYSTEMMODAL | MB_SETFOREGROUND
            # 0x40 | 0x1000 | 0x10000 = 0x11040  → erzwingt Vordergrund
            ctypes.windll.user32.MessageBoxW(
                0,
                "Unten links das Koordinaten Format auf  CSV  umstellen (nicht MARC!) dann kopieren.",
                "OSM Extraktor — Bounding Box",
                0x11040
            )
    threading.Thread(target=csv_hinweis_popup, daemon=True).start()
    webbrowser.open("https://boundingbox.klokantech.com")
    bbox = bbox_eingabe()
    print(f"✅ Bounding Box: {bbox}")

    # ── Ausschnitt benennen ───────────────────────────────────────────
    print()
    print("Namen fuer den Ausschnitt eingeben (wird als Dateiname verwendet):")
    print("Beispiel: muenchen_nord, test, testgebiet_1")
    print()
    while True:
        ausschnitt_name = input("Name: ").strip()
        # Nur erlaubte Zeichen (Buchstaben, Zahlen, Bindestrich, Unterstrich)
        import re
        sauber = re.sub(r"[^\w\-]", "_", ausschnitt_name, flags=re.UNICODE)
        sauber = re.sub(r"_+", "_", sauber).strip("_")
        if sauber:
            if sauber != ausschnitt_name:
                print(f"  Bereinigt zu: '{sauber}'  — OK? (j/n)")
                if input().strip().lower() == "n":
                    continue
            ausschnitt_name = sauber
            break
        print("  Bitte einen Namen eingeben.")
    print(f"✅ Ausschnitt-Name: {ausschnitt_name}")

    # ── Ausschneiden + Konvertieren ───────────────────────────────────
    titel("4/5 · Ausschneiden & Konvertieren")

    region_pbf = WORK_DIR / f"{ausschnitt_name}.osm.pbf"
    region_osm = WORK_DIR / f"{ausschnitt_name}.osm"

    extrahieren(osmium, bbox, pbf_pfad, region_pbf)
    konvertieren(osmium, region_pbf, region_osm)

    # ── Analyse ───────────────────────────────────────────────────────
    titel("5/5 · Analyse & Export")
    df = analysieren(region_osm)

    if df.empty:
        print("HINWEIS: Keine Orte gefunden. Bounding Box prüfen!")
        input("\nEnter zum Beenden...")
        return

    xlsx_name = f"kleinorte_{region_name.replace(' ', '_')}_{ausschnitt_name}.xlsx"
    xlsx_pfad = WORK_DIR / xlsx_name
    xlsx_exportieren(df, xlsx_pfad, region_name, bbox)

    # ── Fertig ────────────────────────────────────────────────────────
    trennlinie("═")
    print(f"\n  FERTIG! {len(df)} Orte exportiert.")
    print(f"\n  Datei: {xlsx_pfad}")
    trennlinie("═")

    # Datei öffnen anbieten
    oeffnen = input("\nDatei jetzt öffnen? (j/n) [j]: ").strip().lower()
    if oeffnen != "n":
        os.startfile(str(xlsx_pfad))

    input("\nEnter zum Beenden...")

if __name__ == "__main__":
    main()